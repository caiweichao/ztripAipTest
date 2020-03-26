import openpyxl

from common.logs import log


class test_case:
    # 初始化测试用例
    def __init__(self):
        self.caseId = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class doExcel:
    def __init__(self, fileName):
        try:
            self.fileName = fileName
            log.info(f'打开测试用例数据文件:{fileName}')
            self.wookBook = openpyxl.load_workbook(filename=fileName)
        except FileNotFoundError as e:
            log.error(f'用例文件: {fileName}不存在请检查\n{e}')

    # 读取指定sheet里的用例存放到testCases中
    def getSheetCase(self, sheetName):
        sheet = self.wookBook[sheetName]
        # 获取sheet的最大行数
        maxRow = sheet.max_row
        # 定义测试用例存放的容器
        testCases = []
        # 跳过标同
        for row in range(2, maxRow + 1):
            # 实例化测试用例容器
            testCase = test_case()
            # 定位testCase中每个内容在sheet中的位置
            testCase.caseId = sheet.cell(row=row, column=1).value
            testCase.title = sheet.cell(row=row, column=2).value
            testCase.url = sheet.cell(row=row, column=3).value
            testCase.data = sheet.cell(row=row, column=4).value
            testCase.method = sheet.cell(row=row, column=5).value
            testCase.expected = sheet.cell(row=row, column=6).value
            # 将用例加入测试用例容器
            testCases.append(testCase)
        return testCases

    # 获取testCase的标题用于处理测试报告
    def getTitle(self, sheetName):
        sheet = self.wookBook[sheetName]
        max_row = sheet.max_row
        titles = []
        for row in range(2, max_row + 1):
            title = sheet.cell(row=row, column=2).value
            titles.append(title)
        return titles

    # 接口测试完毕后将结果写入excel
    '''
    :param sheetName: excel标签的名称
    :param caseId: 执行的case的id
    :param actual: 接口返回的实际值
    :param result: 测试结果
    '''

    def writeExcel(self, sheetName, caseId, actual, result):
        sheet = self.wookBook[sheetName]
        maxRow = sheet.max_row
        for row in range(2, maxRow + 1):
            # 取出每一行的caseId
            _caseId = sheet.cell(row=row, column=1).value
            try:
                # 判断取出caseId是否等于传入的caseId
                if _caseId == caseId:
                    # 写入 actual 和 result
                    sheet.cell(row, 7).value = actual
                    sheet.cell(row, 8).value = result
                    self.wookBook.save(filename=self.fileName)
                    break
            except Exception as e:
                log.error(f'写入excel异常请检查\n{e}')

if __name__ == '__main__':
    pass